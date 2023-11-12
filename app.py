from flask import Flask, render_template, request, send_file, jsonify
from flask_cors import CORS
from flask_restful import Api, Resource
from openpyxl import Workbook
from ultralytics import YOLO
from PIL import Image, ImageDraw
from werkzeug.utils import secure_filename
import io
import os
import cv2
import numpy as np

app = Flask(__name__)
api = Api(app)
CORS(app)


# the code gets an img from React and processes it
class FileUploadPhoto(Resource):
    def post(self):
        file = request.files["file"]
        filename = secure_filename(file.filename)
        file.save(os.path.join("img", "input.png"))
        # Call detect
        detect_photo(file)
        return {"message": "File uploaded successfully"}


api.add_resource(FileUploadPhoto, "/api/upload/photo")


def detect_photo(file):
    print("The detect has begun")
    buf = file
    filename = buf.filename
    boxes = detect_objects_on_image(buf.stream)
    # Write the filename and detection results to a text file
    with open("detection_results.txt", "a") as f:
        f.write(f"File: {filename}\n")
        f.write(f"Detection Results: {boxes}\n\n")
    return jsonify(boxes)


def detect_objects_on_image(buf):
    model = YOLO("best.pt")
    results = model.predict(Image.open(buf))
    result = results[0]
    output = Image.open(buf)
    draw = ImageDraw.Draw(output)
    for box in result.boxes:
        x1, y1, x2, y2 = [
            round(x) for x in box.xyxy[0].tolist()
        ]
        class_id = box.cls[0].item()
        prob = round(box.conf[0].item(), 2)
        label = f"{result.names[class_id]} {prob}"
        draw.rectangle([x1, y1, x2, y2], outline="red", width=2)
        draw.text((x1, y1), label, fill="red")
    # Save photo to photos
    output_path = os.path.join("img", "output.png")
    output.save(output_path)
    # Returning an image as a response to a request
    with open(output_path, "rb") as f:
        output_buf = io.BytesIO(f.read())
    return output_buf


@app.route('/api/img')
def get_image():
    return send_file('img/output.png', mimetype='img/png')


# end of working with photo


# the code gets a video from React and processes it
@app.route('/api/upload/video', methods=['POST'])
def upload():
    file = request.files['file']
    file.save(os.path.join('video', 'video.mp4'))
    return 'Video uploaded successfully'


@app.route('/api/video', methods=['GET'])
def video():
    cap = cv2.VideoCapture('/video/video.mp4')
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out = cv2.VideoWriter('video/output.mp4', fourcc, 20.0, (640, 480))

    while cap.isOpened():
        ret, frame = cap.read()
        if ret:
            # Load the neural network model
            model = cv2.dnn.readNet('best.pt', 'best.pt')
            # Preprocess the frame
            blob = cv2.dnn.blobFromImage(frame, 1 / 255, (416, 416), swapRB=True, crop=False)
            # Set the input to the neural network model
            model.setInput(blob)
            # Forward pass through the neural network model
            output = model.forward()
            # Postprocess the output
            boxes = []
            for detection in output:
                scores = detection[5:]
                class_id = np.argmax(scores)
                confidence = scores[class_id]
                if confidence > 0.5:
                    center_x = int(detection[0] * 416)
                    center_y = int(detection[1] * 416)
                    width = int(detection[2] * 416)
                    height = int(detection[3] * 416)
                    x1 = int(center_x - width / 2)
                    y1 = int(center_y - height / 2)
                    x2 = int(center_x + width / 2)
                    y2 = int(center_y + height / 2)
                    boxes.append([x1, y1, x2, y2, class_id])
            # Draw the boxes on the frame
            for box in boxes:
                x1, y1, x2, y2, class_id = box
                cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.putText(frame, str(class_id), (x1, y1), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2, cv2.LINE_AA)
            # Write the frame to the output video
            out.write(frame)
        else:
            break

    cap.release()
    out.release()

    return send_file('video/output.mp4', mimetype='video/mp4')


# end of working with video


@app.route('/api/send-to-excel')
def send_to_excel():
    # Create new Excel file ( Use this code to write data to an existing Excel file,
    # replacing the string wb = Workbook() with wb = load_workbook('example.xlsx').
    wb = Workbook()
    # Select the active sheet
    ws = wb.active
    # Writing data to cells
    ws['A1'] = 'Hello'
    ws['B1'] = 'World'
    # Saving the file
    wb.save('work.xlsx')


if __name__ == '__main__':
    app.run(debug=True)
